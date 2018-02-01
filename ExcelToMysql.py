# -*- coding: UTF-8 -*-
from openpyxl.reader.excel import load_workbook as lw
import mysql.connector as mc
import sys
import json
import os

def sqlInject():
	config = open(os.path.dirname(os.path.realpath(sys.argv[0]))+"/Setting.json")
	setting = json.load(config)
	filepath = setting["excelPath"]
	dbSetting = setting["db"]
	conn = mc.connect(user=dbSetting["user"], password=dbSetting["password"], host=dbSetting["host"], database=dbSetting["database"])
	cur = conn.cursor()
	update_sql = 'update t_summit set name=%s, xingbie=%s,bumen=%s, zhiwu=%s, zhiji=%s, tel=%s, zhusudidian=%s, fangxing=%s, feiyong=%s, fangjianhao=%s, jieqiaren=%s,jiabinkahao=%s,laichengjiaotong=%s,laichengxinxi=%s,daodashijian=%s,fanchengjiaotong=%s,fanchengxinxi=%s,fanchengshijian=%s,xuanzehuodong=%s,qiandaoxingzhi=%s,laibinxingzhi=%s,jiabinzhuangtai=%s,yuanzhuohuiyi=%s,beizhu1=%s,beizhu2=%s,beizhu3=%s,beizhu4=%s  where id=%s'

	#获取excel文件内容
	wb = lw(filename = filepath)
	ws = wb[(wb.sheetnames[0])]

	#处理数据    
	rows = ws.max_row-2
	columns = ws.max_column+4
	data = []
	daodashijian=""
	fanchengshijian=""
	for rx in range(3, rows):
		for cx in range(1, columns):
			string = str(ws.cell(row=rx, column=cx).value) 
			if string == "None":
				string=None
			data.append(string)

		if data[15]:
			data[15] = data[15].replace("年","/")
			data[15] = data[15].replace("月","/")
			data[15] = data[15].replace("日"," ")
		if data[18]:
			data[18] = data[18].replace("年","/")
			data[18] = data[18].replace("月","/")
			data[18] = data[18].replace("日"," ")

		#执行sql语句
		cur.execute(update_sql, (data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9],data[10],data[11],data[12],data[13],data[14],data[15],data[16],data[17],data[18],data[19],data[20],data[21],data[22],data[23],data[24],data[25],data[26],data[27],data[0]))
		data = []
	conn.commit()  # 提交
	# 关闭两个连接
	print "完成"
	cur.close()    
	conn.close()	

if __name__ == '__main__':
	reload(sys)
	sys.setdefaultencoding('utf-8')
	sqlInject()
